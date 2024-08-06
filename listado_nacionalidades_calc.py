from listado_nacionalidades import *
import leew
from openpyxl import load_workbook # uso https://realpython.com/lessons/reading-excel-sheets/


class MainWindow(QtWidgets.QMainWindow, Ui_listadodenacionalidades):
    def __init__(self,obj, *args,user='', **kwargs):
        QtWidgets.QMainWindow.__init__(self, *args, **kwargs)
        self.setupUi(self)

        self.obj = obj  # para poder habilitar de nuevo el menu de listado de trabajaodes
        #variable para determinar que usuario modificó la informacion
        self.usuario = user
        # variable de control para close
        self.cerrar = 0
        # variable para grabar fechas de modificacion en la BD
        self.dia = QtCore.QDateTime.currentDateTime().toString("dd-MM-yyyy, hh:mm:ss AP")

        self.refresh_tabla()
        self.importar.clicked.connect(self.importar_cmd)

        self.cancelar.clicked.connect(self.close)

        self.guardar.clicked.connect(self.guardar_cmd)

        self.guardar.setDisabled(1)

    def refresh_tabla(self):
        # Llenando las filas
        lista_id = []

        try:  # para atrapar una BD de vacia
            lista_id = leew.consulta_lista('worker.db', 'id', 'listado_nacionalidades', 'id>', '"0"')
            n = len(lista_id)
        except:
            n = 0

        self.tableWidget.setRowCount(n)

        # estas lineas de abajo son para estirar las columnas
        header = self.tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)

        fila = 0
        for i in lista_id:
            i = str(i)
            item = leew.consulta_gen('worker.db', 'id_tss', 'listado_nacionalidades', 'id', f'"{i}"')
            self.tableWidget.setItem(fila, 0, QtWidgets.QTableWidgetItem(str(item)))
            item = leew.consulta_gen('worker.db', 'nacionalidad', 'listado_nacionalidades', 'id', f'"{i}"')
            self.tableWidget.setItem(fila, 1, QtWidgets.QTableWidgetItem(item))
            fila = fila + 1

    def importar_cmd(self):
        # importa un archivo de excel .xlsx
        try:
            ruta_archivo = QtWidgets.QFileDialog.getOpenFileName(self, '', '', '*.xlsx')[0]  # *.png *.svg*.jpg
            #print(type(ruta_archivo))
            if ruta_archivo == '':
                pass # cuando se presiona cancelar en la ventana de de selección de archivos
            else:
                self.tableWidget.setRowCount(0)  # para borrar previamente la tabla antes de importar
                archivo = load_workbook(filename=ruta_archivo)
                cant_filas = archivo.active.max_row - 1 # para restar el encabezado de la hoja
                #print(cant_filas)

                self.tableWidget.setRowCount(int(cant_filas))
                for fila in range(1,cant_filas + 1):# para restar el encabezado de la hoja
                    item = str(archivo.active.cell(row=fila + 1,column=1).value)
                    self.tableWidget.setItem(fila - 1, 0, QtWidgets.QTableWidgetItem(item))
                    item = archivo.active.cell(row=fila + 1,column=2).value
                    self.tableWidget.setItem(fila - 1, 1, QtWidgets.QTableWidgetItem(item))

                self.guardar.setDisabled(0)# ya que se hizo una carga de archivo ahora se puede guardar
        except:
            QtWidgets.QMessageBox.warning(self, "Error", "Ocurrio un error en la operación",
                                          QtWidgets.QMessageBox.Ok)

    def guardar_cmd(self):
        try:
            leew.del_gen_todas_las_filas('worker.db','listado_nacionalidades')
            numero_de_filas = self.tableWidget.rowCount()
            for fila in range(numero_de_filas):
                valores = 'NULL,"' + self.tableWidget.item(fila,0).text() + '","' + self.tableWidget.item(fila,1).text() + \
                          '","' +  self.dia + '","' + self.usuario + '"'
                #print(valores)
                leew.introduce_gen('worker.db','listado_nacionalidades',valores)
            QtWidgets.QMessageBox.information(self, "Atención", "Listado de nacionalidades modificado satisfactoriamente",
                                              QtWidgets.QMessageBox.Ok)
            self.guardar.setDisabled(1) # ya que se guardó ahora se desactiva el boton hasta que se vuelva a importar
            self.cerrar = 1  # para que no pregunte al cerrar
        except:
            QtWidgets.QMessageBox.warning(self, "Error", "Ocurrio un error en la operación",
                                              QtWidgets.QMessageBox.Ok)

    def closeEvent(self, QCloseEvent):

        if self.cerrar == 0:
            reply = QtWidgets.QMessageBox.question(self, 'Advertencia',
                                                   '¿Está usted seguro de salir? Se perderán los datos '
                                                   'no guardados', QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                                                   QtWidgets.QMessageBox.No)

            if reply == QtWidgets.QMessageBox.Yes:

                QCloseEvent.accept()
                self.parentWidget().close()
                self.obj.setDisabled(False)

            else:

                QCloseEvent.ignore()
        else:
            QCloseEvent.accept()
            self.parentWidget().close()
            self.obj.setDisabled(False)
