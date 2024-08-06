from tabla_part_beneficios import *

import leew, ver_recibos, part_beneficios_calc, ordenpago, os, genera_txt_bonificacion

from openpyxl import Workbook

class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):

    def __init__(self,obj,obj2, *args, user='', **kwargs):
        QtWidgets.QMainWindow.__init__(self, *args, **kwargs)

        self.setupUi(self)

        self.user = user # este usuario es para grabarlo en la bd cuando se crea o edita un trbajador

        self.obj = obj # para que las fichas sea subwindows mdi
        self.obj2 = obj2 # para poder habilitar de nuevo el menu de listado de trabajaodes
        #print(self.obj)
        # Carga de data

        self.refresh_tabla()
        self.carga_info_en_tabla_periodos_abiertos()

        # conexiones

        self.cerrar.clicked.connect(self.cerrar_cmd)

        self.tableWidget.clicked.connect(self.cambia_nombre)
        self.gen_recibos.setDisabled(1)
        self.gen_orden_pagos.setDisabled(1)
        self.exportar.setDisabled(1)
        self.txt.setDisabled(1)


        self.correr_nomina.clicked.connect(self.correr_nomina_cmd)
        self.gen_recibos.clicked.connect(self.generar_recibos_cmd)
        self.gen_orden_pagos.clicked.connect(self.generar_orden_de_pago_cmd)
        self.exportar.clicked.connect(self.exportar_a_excel)
        self.txt.clicked.connect(self.genera_txt)

        self.correr_nomina.setFocus() # para obtener el foco
        self.correr_nomina.setDefault(1) # para que se preseleccione y poder dar enter

    def correr_nomina_cmd(self):

        if leew.consulta_lista('worker.db', 'idp', 'periodo_fiscal', 'status', '"ABIERTO" LIMIT 1') == []:
            QtWidgets.QMessageBox.warning(self, "Advertencia",
                                          "No existen periodos fiscales abiertos.",
                                          QtWidgets.QMessageBox.Ok)

        else:
            self.primer_periodo_abierto = leew.consulta_lista('worker.db', 'idp', 'periodo_fiscal', 'status', '"ABIERTO" LIMIT 1')[0]
            self.correr_nomina.setDisabled(True)
            self.obj.addSubWindow(part_beneficios_calc.MainWindow(self.refresh_tabla,self.correr_nomina,self.primer_periodo_abierto, self,user=self.user)).show()

    def carga_info_en_tabla_periodos_abiertos(self):
        # Llenando las filas
        lista_periodos_fiscales_abiertos = leew.consulta_lista('worker.db', 'idp', 'periodo_fiscal', 'status ',
                                                               '"ABIERTO"')
        # print(lista_periodos_fiscales)
        self.tabla_per_fiscal_abiertos.setRowCount(len(lista_periodos_fiscales_abiertos))

        # estas lineas de abajo son para estirar las columnas
        header = self.tabla_per_fiscal_abiertos.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)

        ind = 0
        for periodo_fiscal in lista_periodos_fiscales_abiertos:

            self.tabla_per_fiscal_abiertos.setItem(ind, 0, QtWidgets.QTableWidgetItem(periodo_fiscal))

            ind = ind + 1

    def generar_recibos_cmd(self):

        ver_recibos.ver_recibos(self.nomina, tipo_nomina='beneficios')

    def generar_orden_de_pago_cmd(self):
        ordenpago.imprimir(self.nomina, tipo_nomina='beneficios')

    def refresh_tabla(self):

        # Llenando las filas
        lista_periodos_fiscales = leew.consulta_lista_distintc('worker.db','idp_fiscal','nomina_beneficios','idn >', '0')
        #print(lista_periodos_fiscales)
        self.tableWidget.setRowCount(len(lista_periodos_fiscales))
        lista_periodos_fiscales.reverse() # esto para que la ultima periodo_fiscal salga de primero

        # estas lineas de abajo son para estirar las columnas
        header = self.tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.Stretch)

        ind = 0
        for periodo_fiscal in lista_periodos_fiscales:

            item = leew.consulta_gen('worker.db','fecha_doc','nomina_beneficios','idp_fiscal',f'"{periodo_fiscal}"')
            self.tableWidget.setItem(ind, 0, QtWidgets.QTableWidgetItem(item))

            self.tableWidget.setItem(ind, 1, QtWidgets.QTableWidgetItem(periodo_fiscal))
            item = leew.consulta_gen_sum('worker.db','total_ajustado', 'nomina_beneficios', 'idp_fiscal', f'"{periodo_fiscal}"')
            self.tableWidget.setItem(ind, 2, QtWidgets.QTableWidgetItem(str("{:,.2f}".format(round(item,2)))))
            self.tableWidget.item(ind,2).setTextAlignment(2999)
            item = leew.consulta_gen('worker.db','nota','nomina_beneficios','idp_fiscal',f'"{periodo_fiscal}"')
            self.tableWidget.setItem(ind, 3, QtWidgets.QTableWidgetItem(item))
            ind = ind + 1
        self.carga_info_en_tabla_periodos_abiertos()

    def cambia_nombre(self):

        self.gen_recibos.setDisabled(0)
        self.gen_orden_pagos.setDisabled(0)
        self.exportar.setDisabled(0)
        self.txt.setDisabled(0)
        for currentQTableWidgetItem in self.tableWidget.selectedItems():
            self.nomina = self.tableWidget.item(currentQTableWidgetItem.row(), 1).text()
            self.gen_recibos.setText("Recibos: \n" + self.nomina)

            self.gen_orden_pagos.setText("Orden de pago: \n" + self.nomina)

            self.exportar.setText("Exportar: \n" + self.nomina)
            self.txt.setText(f"TXT: \n{self.nomina}")

    def exportar_a_excel(self):
        filas_tabla = leew.exporta_filas_tabla('worker.db', f'nomina_beneficios WHERE idp_fiscal = "{self.nomina}"')
        #print(filas_tabla)
        cabecera = ['idn', 'Per Fiscal','Utilidades monto','Porcentaje','Total a repartir', 'Total segun codigo',
                    'Factor de ajuste', 'Total Ajustado','Id','Trabajador','Ingreso','Egreso','Salario mensual',
                    'Salario diario','Tiempo trabajado', 'Cant días','Monto bruto','Monto ajustado','Ret INFOTEP',
                    'Ret ISLR', 'Monto pagado','Nota', 'Fecha de computo']

        wb = Workbook() # crea el libro de excell
        ws = wb.active # activa la hoja de excel que se crea por default
        # para la cabecera
        c = 1
        for nombre in cabecera:
            ws.cell(row=1, column= c, value=nombre)
            c = c + 1
        f = 2
        for fila in filas_tabla:
            c = 1
            for item in fila[0:23]:
                ws.cell(row=f, column=c, value=item)
                c = c + 1
            f = f + 1


        ruta_a_exportar = \
            QtWidgets.QFileDialog.getSaveFileName(self, '', f'C:\\Users\\{os.environ.get("USERNAME")}', '*.xlsx')[0]
        ruta_a_exportar = os.path.abspath(ruta_a_exportar)
        #print(ruta_a_exportar)

        if ruta_a_exportar == os.path.dirname(os.path.abspath(
                __file__)):  # esto sucede cuando se presiona el boton cancelar en la ventana de seleccion de archivo
            pass

        else:
            # print(ruta_de_instalacio_prog)

            reply = QtWidgets.QMessageBox.question(self, 'Advertencia',
                                                   '¿Está usted seguro de exportar a MS Excel(R) la tabla de pago de la participación en los beneficios?'
                                                   , QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                                                   QtWidgets.QMessageBox.No)

            if reply == QtWidgets.QMessageBox.Yes:
                try:

                    wb.save(ruta_a_exportar)
                    guardado = QtWidgets.QMessageBox.warning(self, "Operación exitosa", "La exportación se realizó correctamente",
                                                  QtWidgets.QMessageBox.Ok)
                    os.startfile(ruta_a_exportar)

                except Exception as exception:
                    QtWidgets.QMessageBox.warning(self, "Error fatal", f"{type(exception).__name__}, {exception}",
                                                  QtWidgets.QMessageBox.Ok)

    def genera_txt(self):
        genera_txt_bonificacion.generar_txt(self.nomina)

    def cerrar_cmd(self):
        self.close()
        self.parentWidget().close()

    def closeEvent(self, QCloseEvent):
        # esta funcion es para que el boton cerrar y la X de la ventana reactiven el menu de listado de trabajadores
        self.close()
        self.parentWidget().close()
        self.obj2.setDisabled(False)
