import leew
from part_beneficios import *

import trat_fecha, os
from datetime import date
from openpyxl import Workbook

class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):

    def __init__(self,cmd,obj, primer_periodo_abierto, *args,user='', **kwargs):
        QtWidgets.QMainWindow.__init__(self, *args, **kwargs)

        self.setupUi(self)

        self.ya_se_puede_cerrar = False

        self.MMAAAA = date.today().strftime("%m%Y")

        self.user = user # este usuario es para grabarlo en la bd cuando se crea o edita un trabajador
        self.refrescar_tabla_nominas = cmd # para que las fichas sea subwindows mdi

        self.obj = obj
        self.primer_periodo_abierto = primer_periodo_abierto

        self.periodo_fiscal_base.setText(self.primer_periodo_abierto)

        self.dias_por_mes = leew.consulta_gen('worker.db', 'dias_mes', 'beneficios', 'status', '"Vigente"')
        self.tasa_infotep_trabajador = leew.consulta_gen('worker.db', 'infotep_trab', 'legales', 'status', '"Vigente"')
        self.fecha_cierre_fiscal = leew.consulta_gen('worker.db','fecha_f', 'periodo_fiscal', 'idp', f'"{self.primer_periodo_abierto}"')
        self.fecha_cierre_fiscal = [int(n) for n in self.fecha_cierre_fiscal.split('-')] # una lista de enteros [anio, mes, dia]
        #print(self.fecha_cierre_fiscal)
        self.cargar_data_en_tabla()
        self.correr.setDisabled(True)

        # conexiones

        self.cancelar.clicked.connect(self.cerrar_cmd)
        self.monto_utilidades.valueChanged.connect(self.cargar_data_en_tabla)
        self.porc_repartir.valueChanged.connect(self.cargar_data_en_tabla)
        self.rango1.valueChanged.connect(self.cargar_data_en_tabla)
        self.rango2.valueChanged.connect(self.cargar_data_en_tabla)
        self.anio_cambio.valueChanged.connect(self.cargar_data_en_tabla)
        self.exportar_xls.clicked.connect(self.exportar_a_excel)
        self.nota.textChanged.connect(self.habilitador_correr)

        self.correr.clicked.connect(self.correr_nomina_cmd)

    def cargar_data_en_tabla(self):

        lista_mmaaaa_por_revisar = leew.consulta_gen('worker.db', 'lista_mmaaaa', 'periodo_fiscal', 'idp',
                                                     f'"{self.primer_periodo_abierto}"')
        lista_mmaaaa_por_revisar = lista_mmaaaa_por_revisar.split(',')
        #print(lista_mmaaaa_por_revisar)
        lista_de_trabajadores_en_el_periodo_fiscal = []
        for mmaaaa in lista_mmaaaa_por_revisar:
            id = leew.consulta_lista('worker.db', 'ID_TRABAJADOR', 'nomina', 'MMAAAA', f'"{mmaaaa}"')
            lista_de_trabajadores_en_el_periodo_fiscal = lista_de_trabajadores_en_el_periodo_fiscal + id

        set_de_trabajadores_en_el_periodo_fiscal = set(lista_de_trabajadores_en_el_periodo_fiscal) # set para eliminar repetidos
        #print(set_de_trabajadores_en_el_periodo_fiscal)
        self.tableWidget.setRowCount(len(set_de_trabajadores_en_el_periodo_fiscal))

        # estas lineas de abajo son para estirar las columnas
        header = self.tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(6, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(7, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(8, QtWidgets.QHeaderView.ResizeToContents)

        tbj = 0
        total_segun_codigo = 0
        for trabajador in set_de_trabajadores_en_el_periodo_fiscal:
            trabajador = str(trabajador)
            #id
            self.tableWidget.setItem(tbj, 0, QtWidgets.QTableWidgetItem(trabajador))
            nombre = leew.consulta_gen('worker.db', 'Nombre', 'info', 'id', f'"{trabajador}"')
            apellido = leew.consulta_gen('worker.db', 'Apellido', 'info', 'id', f'"{trabajador}"')
            nombre_completo = f"{nombre} {apellido}"
            self.tableWidget.setItem(tbj, 1, QtWidgets.QTableWidgetItem(nombre_completo))
            fecha_ingreso = leew.consulta_gen('worker.db', 'Fecha_ingreso', 'info', 'id', f'"{trabajador}"')
            self.tableWidget.setItem(tbj, 2, QtWidgets.QTableWidgetItem(fecha_ingreso))
            fecha_egreso = leew.consulta_gen('worker.db', 'fecha_egreso', 'liquidacion', 'id_trab', f'"{trabajador}"')
            if fecha_egreso == None:
                fecha_egreso = 'Activo'
            self.tableWidget.setItem(tbj, 3, QtWidgets.QTableWidgetItem(fecha_egreso))
            # salario diario
            salario = 0
            inasis = 0
            comisiones = 0
            otras_rem = 0
            otras_rem_no_sal = 0
            vacaciones = 0
            vac_pago = 0
            for mmaaaa in lista_mmaaaa_por_revisar:
                salarioBD = leew.consulta_gen_sum('worker.db', 'SALARIO_QUINCENA', 'nomina', 'MMAAAA', f'"{mmaaaa}" AND ID_TRABAJADOR = "{trabajador}"')
                if salarioBD == None:
                    salarioBD = 0
                salario = salario + salarioBD
                inasisBD = leew.consulta_gen_sum('worker.db', 'INASIS', 'nomina', 'MMAAAA', f'"{mmaaaa}" AND ID_TRABAJADOR = "{trabajador}"')
                if inasisBD == None:
                    inasisBD = 0
                inasis = inasis + inasisBD
                comisionesBD = leew.consulta_gen_sum('worker.db', 'COMISIONES', 'nomina', 'MMAAAA', f'"{mmaaaa}" AND ID_TRABAJADOR = "{trabajador}"')
                if comisionesBD == None:
                    comisionesBD = 0
                comisiones = comisiones + comisionesBD
                otras_remBD = leew.consulta_gen_sum('worker.db', 'OTRAS_REMUN', 'nomina', 'MMAAAA', f'"{mmaaaa}" AND ID_TRABAJADOR = "{trabajador}"')
                if otras_remBD == None:
                    otras_remBD = 0
                otras_rem = otras_rem + otras_remBD
                otras_rem_no_salBD = leew.consulta_gen_sum('worker.db', 'OTRAS_REMUN_NO_SALARIALES', 'nomina', 'MMAAAA',
                                                    f'"{mmaaaa}" AND ID_TRABAJADOR = "{trabajador}"')
                if otras_rem_no_salBD == None:
                    otras_rem_no_salBD = 0
                otras_rem_no_sal = otras_rem_no_sal + otras_rem_no_salBD
                vacacionesBD = leew.consulta_gen_sum('worker.db', 'FRACCION_VAC', 'nomina', 'MMAAAA', f'"{mmaaaa}" AND ID_TRABAJADOR = "{trabajador}"')
                if vacacionesBD == None:
                    vacacionesBD = 0
                vacaciones = vacaciones + vacacionesBD
                vac_pagoBD = leew.consulta_gen_sum('worker.db', 'PAGO_DE_VAC', 'nomina', 'MMAAAA',
                                                     f'"{mmaaaa}" AND ID_TRABAJADOR = "{trabajador}"')
                if vac_pagoBD == None:
                    vac_pagoBD = 0
                vac_pago = vac_pago + vac_pagoBD
            salario_mensual = (salario - inasis + comisiones + otras_rem - otras_rem_no_sal + vacaciones + vac_pago) / 12
            self.tableWidget.setItem(tbj, 4, QtWidgets.QTableWidgetItem("RD${:,.2f}".format(salario_mensual)))
            salario_diario = salario_mensual / self.dias_por_mes
            self.tableWidget.setItem(tbj, 5, QtWidgets.QTableWidgetItem("RD${:,.2f}".format(salario_diario)))
            fecha_ingreso = [int(n) for n in fecha_ingreso.split('-')] #ahora como lista de enteros(int) [dia, mes, anio]
            #print(fecha_ingreso)
            if fecha_egreso == 'Activo':
                tiempo_trab = trat_fecha.calc_dif_fecha(date(fecha_ingreso[2], fecha_ingreso[1], fecha_ingreso[0]),
                                                              date(self.fecha_cierre_fiscal[2],self.fecha_cierre_fiscal[1], self.fecha_cierre_fiscal[0]))
            else:# el trabajador ya no trabaja, egresó
                fecha_egreso_lista = [int(n) for n in fecha_egreso.split('-')]
                tiempo_trab = trat_fecha.calc_dif_fecha(date(fecha_ingreso[2], fecha_ingreso[1], fecha_ingreso[0]),
                                                        date(fecha_egreso_lista[2], fecha_egreso_lista[1],
                                                             fecha_egreso_lista[0]))
            tiempo_trabajando_formato = f"{tiempo_trab[0]} años, {tiempo_trab[1]} meses, {tiempo_trab[2]} días"
            self.tableWidget.setItem(tbj, 6, QtWidgets.QTableWidgetItem(tiempo_trabajando_formato))
            if tiempo_trab[0] < self.anio_cambio.value():
                cant_dias = self.rango1.text()
            else:
                cant_dias = self.rango2.text()
            self.tableWidget.setItem(tbj, 7, QtWidgets.QTableWidgetItem(cant_dias))
            monto_bruto = int(cant_dias) * salario_diario
            self.tableWidget.setItem(tbj, 8, QtWidgets.QTableWidgetItem("RD${:,.2f}".format(monto_bruto)))
            total_segun_codigo = total_segun_codigo + monto_bruto # aqui sumo el monto bruto de cada trabajador
            tbj = tbj + 1
        self.total_segun_codigo.setValue(total_segun_codigo)
        self.total_a_repartir.setValue(self.porc_repartir.value() * self.monto_utilidades.value() / 100)
        #print(self.total_a_repartir.value())
        if self.total_segun_codigo.value() > self.total_a_repartir.value():
            self.factor_de_ajuste.setValue(self.total_a_repartir.value() / self.total_segun_codigo.value())
        else:
            self.factor_de_ajuste.setValue(1.0)

        total_ajustado = 0
        for fila in range(self.tableWidget.rowCount()):
            monto_ajustado = self.factor_de_ajuste.value() * float(self.tableWidget.item(fila, 8).text().replace(',','').replace('RD$',''))
            total_ajustado = total_ajustado + monto_ajustado
            self.tableWidget.setItem(fila, 9, QtWidgets.QTableWidgetItem("RD${:,.2f}".format(monto_ajustado)))
            ret_infotep = monto_ajustado * self.tasa_infotep_trabajador
            self.tableWidget.setItem(fila, 10, QtWidgets.QTableWidgetItem(
                "RD${:,.2f}".format(ret_infotep)))
            id_trabajador = self.tableWidget.item(fila,0).text()
            retencion_isr = self.retencion_ISLR(id_trabajador, monto_ajustado)
            self.tableWidget.setItem(fila, 11, QtWidgets.QTableWidgetItem(
                "RD${:,.2f}".format(retencion_isr)))
            self.tableWidget.setItem(fila, 12, QtWidgets.QTableWidgetItem(
                "RD${:,.2f}".format(monto_ajustado - retencion_isr - ret_infotep)))
        self.total_ajustado.setValue(total_ajustado)

    def retencion_ISLR(self, id_trabajador, bonificacion):
        salario_base_isr_mes = leew.consulta_gen_sum('worker.db', 'ISLR_BASE','nomina','ID_TRABAJADOR',
                                                     f'{id_trabajador} AND MMAAAA="{self.MMAAAA}"')
        if salario_base_isr_mes == None:
            salario_base_isr_mes = 0
        retencio_isr_mes = leew.consulta_gen_sum('worker.db', 'ISLR_RETENCION','nomina','ID_TRABAJADOR',
                                                     f'{id_trabajador} AND MMAAAA="{self.MMAAAA}"')
        if retencio_isr_mes == None:
            retencio_isr_mes = 0
        salario_base_isr_egreso_vac = leew.consulta_gen_sum('worker.db', 'vac_monto','liquidacion','id_trab',
                                                     f'{id_trabajador} AND MMAAAA="{self.MMAAAA}"')
        if salario_base_isr_egreso_vac == None:
            salario_base_isr_egreso_vac = 0
        salario_base_isr_egreso_sal_nav = leew.consulta_gen_sum('worker.db', 'parte_grav_sal_nav', 'liquidacion', 'id_trab',
                                                            f'{id_trabajador} AND MMAAAA="{self.MMAAAA}"')
        if salario_base_isr_egreso_sal_nav == None:
            salario_base_isr_egreso_sal_nav = 0
        ret_islr_egreso = leew.consulta_gen_sum('worker.db', 'islr_ret', 'liquidacion', 'id_trab',
                                                            f'{id_trabajador} AND MMAAAA="{self.MMAAAA}"')
        if ret_islr_egreso == None:
            ret_islr_egreso = 0

        return leew.calculo_retencion_islr(salario_base_isr_mes + salario_base_isr_egreso_vac
                                           + salario_base_isr_egreso_sal_nav + bonificacion) - retencio_isr_mes -\
                                            ret_islr_egreso

    def exportar_a_excel(self):
        wb = Workbook() # crea el libro de excell
        ws = wb.active # activa la hoja de excel que se crea por default
        for columna in range(self.tableWidget.horizontalHeader().count()):
            ws.cell(row=1, column=columna + 1, value=self.tableWidget.horizontalHeaderItem(columna).text())
            for fila in range(self.tableWidget.rowCount()):
                ws.cell(row = fila + 2, column= columna + 1, value=self.tableWidget.item(fila, columna).text().replace('RD$',''))

        ruta_a_exportar = \
            QtWidgets.QFileDialog.getSaveFileName(self, '', f'C:\\Users\\{os.environ.get("USERNAME")}', '*.xlsx')[0]
        ruta_a_exportar = os.path.abspath(ruta_a_exportar)
        #print(ruta_a_exportar)

        if ruta_a_exportar == os.path.dirname(os.path.abspath(
                __file__)):  # esto sucede cuando se presiona el boton cancelar en la ventana de seleccion de archivo
            pass

        else:
            # print(ruta_de_instalacio_prog)

            reply = QtWidgets.QMessageBox.question(self, 'Advertencia',
                                                   '¿Está usted seguro de exportar a MS Excel(R) la tabla de pago de la participación en los beneficios?'
                                                   , QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                                                   QtWidgets.QMessageBox.No)

            if reply == QtWidgets.QMessageBox.Yes:
                try:

                    wb.save(ruta_a_exportar)
                    guardado = QtWidgets.QMessageBox.warning(self, "Operación exitosa", "La exportación se realizó correctamente",
                                                  QtWidgets.QMessageBox.Ok)
                    os.startfile(ruta_a_exportar)

                except Exception as exception:
                    QtWidgets.QMessageBox.warning(self, "Error fatal", f"{type(exception).__name__}, {exception}",
                                                  QtWidgets.QMessageBox.Ok)

    def correr_nomina_cmd(self):

        reply = QtWidgets.QMessageBox.question(self, 'Para continuar',
                                               f'¿Está usted seguro correr la nómina correspondiente al periodo: {self.primer_periodo_abierto}?'
                                               , QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                                               QtWidgets.QMessageBox.No)

        if reply == QtWidgets.QMessageBox.Yes:

            leew.backup_bd('nomina_part_beneficios')
            for linea in range(self.tableWidget.rowCount()):
                entrada = f'NULL, "{self.primer_periodo_abierto}", "{self.monto_utilidades.value()}", ' \
                          f'"{self.porc_repartir.value()}", "{self.total_a_repartir.value()}", "{self.total_segun_codigo.value()}",' \
                          f'"{self.factor_de_ajuste.value()}", "{self.total_ajustado.value()}",' \
                          f'"{self.tableWidget.item(linea, 0).text()}", "{self.tableWidget.item(linea, 1).text()}",' \
                          f'"{self.tableWidget.item(linea, 2).text()}", "{self.tableWidget.item(linea, 3).text()}",' \
                          f'"{self.tableWidget.item(linea, 4).text()}", "{self.tableWidget.item(linea, 5).text().replace("RD$","").replace(",","")}",' \
                          f'"{self.tableWidget.item(linea, 6).text().replace("RD$","").replace(",","")}", "{self.tableWidget.item(linea, 7).text()}",' \
                          f'"{self.tableWidget.item(linea, 8).text().replace("RD$","").replace(",","")}", "{self.tableWidget.item(linea, 9).text().replace("RD$","").replace(",","")}",' \
                          f'"{self.tableWidget.item(linea, 10).text().replace("RD$","").replace(",","")}","{self.tableWidget.item(linea, 11).text().replace("RD$","").replace(",","")}",' \
                          f'"{self.tableWidget.item(linea, 12).text().replace("RD$","").replace(",","")}",' \
                          f'"{self.nota.text()}", "{date.today().strftime("%d-%m-%Y")}", "{self.MMAAAA}", "{self.user}",' \
                          f'NULL, NULL, NULL, NULL'
                leew.introduce_gen('worker.db', 'nomina_beneficios', entrada)

            # cierre de periodo fiscal
            leew.update_gen('worker.db','periodo_fiscal', 'status', f'"CERRADO" WHERE idp = "{self.primer_periodo_abierto}"')

            # imprimir recibos y orden de pago

            QtWidgets.QMessageBox.information(self,"Aviso", "Nómina procesada satisfactoriamente")
            try:
                self.refrescar_tabla_nominas()
            except:
                pass

            self.ya_se_puede_cerrar = True
            self.close()
            try:
                self.obj.setDisabled(0)
            except:
                pass
        else:
           pass

    def habilitador_correr(self):
        if self.nota.text() != '':
            self.correr.setDisabled(False)

        else:
            self.correr.setDisabled(True)

    def cerrar_cmd(self):
        self.close()

    def closeEvent(self, QCloseEvent):
        if self.ya_se_puede_cerrar == False:
            reply = QtWidgets.QMessageBox.question(self, 'Advertencia',
                                                   '¿Está usted seguro de salir? Se perderán los datos '
                                                   'no guardados', QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                                                   QtWidgets.QMessageBox.No)

            if reply == QtWidgets.QMessageBox.Yes:

                QCloseEvent.accept()
                self.parentWidget().close()
                try:
                    self.obj.setDisabled(0)
                except:
                    pass
            else:
                QCloseEvent.ignore()
        else:
            QCloseEvent.accept()
            self.parentWidget().close()
            try:
                self.obj.setDisabled(0)
            except:
                pass